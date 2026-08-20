from __future__ import annotations

import hashlib
import html
import re
import shutil
import typing
from pathlib import Path
from uuid import uuid4

from jinja2 import Environment, StrictUndefined, select_autoescape
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.core.config import settings
from app.core.time import beijing_now, format_beijing
from app.models import (
    Artifact,
    ConfigurationCaptureSnapshot,
    Resource,
    RunComparison,
    RunStep,
    TestRun,
)
from app.services.order_configs import OrderConfigError, parse_xml


REPORT_ARTIFACT_TYPES = frozenset({"web_report", "excel_report", "pdf_report"})
SENSITIVE_ARTIFACT_TYPES = REPORT_ARTIFACT_TYPES | frozenset({"order_config_xml"})
REPORT_VERSION_RE = re.compile(r"^report-v(\d+)\.(?:html|xlsx|pdf)$")
METRIC_ORDER = (
    "average", "maximum", "minimum", "median", "stddev", "sample_count",
    "excluded_above_limit", "excluded_negative", "excluded_invalid",
    "p0_1", "p0_5", "p1", "p5", "p10", "p25", "p50", "p75", "p90",
    "p95", "p99", "p99_5", "p99_9",
)
METRIC_LABELS = {
    "average": "Avg",
    "maximum": "Max",
    "minimum": "Min",
    "median": "Md",
    "stddev": "Std",
    "sample_count": "Cnt",
    "excluded_above_limit": "超上限剔除",
    "excluded_negative": "负数剔除",
    "excluded_invalid": "无效剔除",
}


REPORT_TEMPLATE = """<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<title>OpenSLT 测速报告 {{ report.run_number }}</title>
<style>
@page{size:A4;margin:15mm 13mm 16mm}*{box-sizing:border-box}body{margin:0;color:#18262d;font:12px/1.55 Arial,"Microsoft YaHei",sans-serif}h1{margin:0;color:#126b61;font-size:24px}h2{margin:24px 0 9px;padding-bottom:5px;border-bottom:2px solid #277f74;font-size:16px}h3{margin:16px 0 7px;font-size:13px}.muted{color:#67777f}.meta{margin:4px 0 18px}.empty{padding:12px;border:1px dashed #b7c4c9;color:#6e7d84}.page-break{break-before:page}.table-wrap{max-width:100%;overflow:auto}table{width:100%;border-collapse:collapse;table-layout:fixed}th,td{padding:6px 7px;border:1px solid #9eacb2;text-align:left;vertical-align:top;overflow-wrap:anywhere}th{background:#e8f1ef;font-weight:700}.key{width:22%}.number{text-align:right;font-variant-numeric:tabular-nums}.mono{font-family:"Cascadia Mono",Consolas,monospace;font-size:10px}pre{margin:0;padding:10px;border:1px solid #aab6bb;background:#f7f9f9;white-space:pre-wrap;overflow-wrap:anywhere;font:9px/1.45 "Cascadia Mono",Consolas,monospace}.section-note{margin:5px 0 8px;color:#67777f;font-size:10px}@media print{.table-wrap{overflow:visible}.metrics{break-before:page;page:landscape}h2,h3{break-after:avoid}table,pre{break-inside:auto}tr{break-inside:avoid}}
</style>
</head>
<body>
<h1>OpenSLT 测速报告</h1>
<p class="meta muted">运行编号：{{ report.run_number }}　报告版本：v{{ "%03d"|format(report.report_version) }}　生成时间：{{ report.generated_at }}</p>
<table><tbody>
<tr><th class="key">方案</th><td>{{ report.plan.name or '-' }}</td><th class="key">场景</th><td>{{ report.scenario.name or '-' }}</td></tr>
<tr><th>业务</th><td>{{ report.business_code }}</td><th>运行状态</th><td>{{ report.status }}</td></tr>
<tr><th>开始时间</th><td>{{ report.started_at }}</td><th>结束时间</th><td>{{ report.finished_at }}</td></tr>
</tbody></table>

<h2>服务器配置</h2>
{% if report.servers %}{% for source in report.servers %}
<h3>{{ source.step_name }} / {{ source.resource_name }}（{{ source.resource_type }}）</h3>
<p class="section-note">主机：{{ source.host or '-' }}　采集状态：{{ source.status }}</p>
<table><thead><tr><th>配置项</th><th>采集值</th><th>状态</th></tr></thead><tbody>{% for item in source["items"] %}<tr><td>{{ item.label }}</td><td class="mono">{{ item.value }}</td><td>{{ item.status }}</td></tr>{% endfor %}</tbody></table>
{% endfor %}{% else %}<div class="empty">无服务器配置数据</div>{% endif %}

<h2>数据库配置</h2>
{% if report.databases %}{% for source in report.databases %}
<h3>{{ source.step_name }} / {{ source.database_name or '-' }}</h3>
<table><thead><tr><th>配置键</th><th>采集值</th><th>状态</th></tr></thead><tbody>{% for item in source["items"] %}<tr><td class="mono">{{ item.label }}</td><td class="mono">{{ item.value }}</td><td>{{ item.status }}</td></tr>{% endfor %}</tbody></table>
{% endfor %}{% else %}<div class="empty">无数据库配置数据</div>{% endif %}

<h2>发单 XML 配置</h2>
{% if report.orders %}{% for order in report.orders %}
<h3>{{ order.step_name }} / {{ order.filename }}</h3><p class="section-note mono">SHA-256：{{ order.checksum }}</p>
<table><thead><tr><th>配置路径</th><th>配置项</th><th>值</th></tr></thead><tbody>{% for row in order["rows"] %}<tr><td class="mono">{{ row.path }}</td><td>{{ row.label }}</td><td class="mono">{{ row.value }}</td></tr>{% endfor %}</tbody></table>
{% endfor %}{% else %}<div class="empty">无发单 XML 配置数据</div>{% endif %}

<div class="metrics"><h2>测速指标</h2>
{% for table in report.statistics %}<h3>{{ table.step_name }}</h3><p class="section-note">统计脚本：{{ table.script_filename or '-' }}　单位：ns　异常上限：{{ table.max_latency_ns if table.max_latency_ns is not none else '-' }} ns</p>
<div class="table-wrap"><table><thead><tr><th>指标</th>{% for column in table["columns"] %}<th>{{ column }}</th>{% endfor %}</tr></thead><tbody>{% for row in table["rows"] %}<tr><td>{{ row.label }}</td>{% for value in row["values"] %}<td class="number">{{ value }}</td>{% endfor %}</tr>{% endfor %}</tbody></table></div>
{% else %}<div class="empty">无测速统计数据</div>{% endfor %}</div>

{% if report.comparison %}<div class="metrics"><h2>运行对比</h2>
<p class="section-note">当前运行：{{ report.comparison.target_run_number }}　基线运行：{{ report.comparison.baseline_run_number }}　快照保存时间：{{ report.comparison.saved_at }}　可比性：{{ report.comparison.compatibility }}</p>
{% if report.comparison.warnings %}<p class="section-note">可比性说明：{{ report.comparison.warnings|join('；') }}</p>{% endif %}
<div class="table-wrap"><table><thead><tr><th>指标</th><th>节点 / 数据源</th><th>基线值</th><th>当前值</th><th>绝对变化</th><th>变化率</th><th>判断</th></tr></thead><tbody>{% for row in report.comparison.rows %}<tr><td>{{ row.metric_label }}</td><td>{{ row.step_name }} / {{ row.source_file }}</td><td class="number">{{ row.baseline_display }}</td><td class="number">{{ row.target_display }}</td><td class="number">{{ row.delta_display }}</td><td class="number">{{ row.percentage_display }}</td><td>{{ row.assessment_text }}</td></tr>{% endfor %}</tbody></table></div>
</div>{% endif %}

<h2>步骤时间线</h2><table><thead><tr><th>顺序</th><th>步骤</th><th>类型</th><th>状态</th><th>耗时(ms)</th><th>错误</th></tr></thead><tbody>{% for step in report.steps %}<tr><td>{{ step.position }}</td><td>{{ step.name }}</td><td>{{ step.node_type }}</td><td>{{ step.status }}</td><td class="number">{{ step.duration_ms }}</td><td>{{ step.error }}</td></tr>{% endfor %}</tbody></table>
<h2>最终结论</h2><table><tbody><tr><th class="key">结论</th><td>{{ report.verdict.final_result }}</td></tr><tr><th>问题说明</th><td>{{ report.verdict.issue_description }}</td></tr><tr><th>备注</th><td>{{ report.verdict.notes }}</td></tr></tbody></table>

{% if report.orders %}<div class="page-break"><h2>发单 XML 原文附录</h2>{% for order in report.orders %}<h3>{{ order.step_name }} / {{ order.filename }}</h3><pre>{{ order.raw_xml }}</pre>{% endfor %}</div>{% endif %}
</body></html>"""


def _resource_snapshots(run: TestRun) -> dict[int, dict[str, typing.Any]]:
    resources = (run.config_snapshot or {}).get("resources") or []
    return {
        int(item["id"]): dict(item)
        for item in resources
        if isinstance(item, dict) and item.get("id") is not None
    }


def _capture_sections(
    db: Session, run: TestRun, steps: list[RunStep]
) -> tuple[list[dict[str, typing.Any]], list[dict[str, typing.Any]]]:
    step_by_id = {step.id: step for step in steps}
    step_ids = list(step_by_id)
    if not step_ids:
        return [], []
    snapshots = list(db.scalars(
        select(ConfigurationCaptureSnapshot)
        .where(
            ConfigurationCaptureSnapshot.run_id == run.id,
            ConfigurationCaptureSnapshot.run_step_id.in_(step_ids),
            ConfigurationCaptureSnapshot.scope == "run",
        )
        .options(selectinload(ConfigurationCaptureSnapshot.items))
        .order_by(ConfigurationCaptureSnapshot.run_step_id, ConfigurationCaptureSnapshot.id)
    ).all())
    resource_details = _resource_snapshots(run)
    servers: list[dict[str, typing.Any]] = []
    databases: list[dict[str, typing.Any]] = []
    for snapshot in snapshots:
        step = step_by_id.get(snapshot.run_step_id)
        if step is None:
            continue
        items = [
            {
                "key": item.item_key,
                "label": item.item_label,
                "value": item.value_text or item.error_message or "-",
                "status": item.status,
            }
            for item in snapshot.items
        ]
        if snapshot.source_type == "server":
            resource = resource_details.get(snapshot.resource_id, {})
            servers.append({
                "step_name": step.name,
                "resource_id": snapshot.resource_id,
                "resource_name": resource.get("name") or f"资源 {snapshot.resource_id}",
                "resource_type": resource.get("type") or "server",
                "host": resource.get("host") or "",
                "status": snapshot.status,
                "items": items,
            })
        elif snapshot.source_type == "database":
            databases.append({
                "step_name": step.name,
                "database_name": snapshot.database_name or "",
                "status": snapshot.status,
                "items": items,
            })
    return servers, databases


def _flatten_xml(content: str) -> list[dict[str, str]]:
    try:
        _, document = parse_xml(content)
    except OrderConfigError:
        return [{"path": "/", "label": "XML 原文", "value": "无法解析结构化配置"}]
    rows: list[dict[str, str]] = []

    def visit(node: dict[str, typing.Any], parents: list[str]) -> None:
        name = str(node.get("name") or "node")
        attributes = {
            str(item.get("name")): str(item.get("value") or "")
            for item in node.get("attributes") or []
        }
        identifier = attributes.get("id")
        component = f"{name}[{identifier}]" if identifier else name
        path = parents + [component]
        if "value" in attributes:
            rows.append({
                "path": "/" + "/".join(path),
                "label": attributes.get("disp") or name,
                "value": attributes["value"],
            })
        for attr_name, attr_value in attributes.items():
            if attr_name not in {"id", "disp", "value"}:
                rows.append({
                    "path": "/" + "/".join(path),
                    "label": f"@{attr_name}",
                    "value": attr_value,
                })
        text_value = "".join(
            str(child.get("text") or "")
            for child in node.get("children") or []
            if child.get("type") in {"text", "cdata"}
        ).strip()
        if text_value:
            rows.append({"path": "/" + "/".join(path), "label": "文本", "value": text_value})
        for child in node.get("children") or []:
            if child.get("type") == "element":
                visit(child, path)

    visit(document, [])
    return rows or [{"path": "/", "label": "配置", "value": "无可展示的键值项"}]


def _order_sections(
    db: Session, run: TestRun, steps: list[RunStep]
) -> list[dict[str, typing.Any]]:
    order_steps = [step for step in steps if step.node_type == "order_preparation"]
    if not order_steps:
        return []
    artifacts = list(db.scalars(
        select(Artifact).where(
            Artifact.run_id == run.id,
            Artifact.step_id.in_([step.id for step in order_steps]),
            Artifact.artifact_type == "order_config_xml",
        ).order_by(Artifact.id)
    ).all())
    by_step = {artifact.step_id: artifact for artifact in artifacts}
    sections: list[dict[str, typing.Any]] = []
    for step in order_steps:
        artifact = by_step.get(step.id)
        if artifact is None:
            continue
        path = Path(artifact.path)
        if not path.is_file():
            continue
        data = path.read_bytes()
        if hashlib.sha256(data).hexdigest() != artifact.checksum:
            continue
        content = data.decode("utf-8")
        sections.append({
            "step_id": step.id,
            "step_name": step.name,
            "filename": artifact.name,
            "checksum": artifact.checksum,
            "artifact_id": artifact.id,
            "rows": _flatten_xml(content),
            "raw_xml": content,
        })
    return sections


def _format_number(value: typing.Any) -> str:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return "-"
    if number.is_integer():
        return str(int(number))
    return f"{number:.3f}".rstrip("0").rstrip(".")


def _statistics_table(step: RunStep) -> typing.Optional[dict[str, typing.Any]]:
    summary = step.result_summary or {}
    results = summary.get("statistics_results")
    if not isinstance(results, list) or not results:
        return None
    columns: list[str] = []
    values_by_result: list[dict[str, typing.Any]] = []
    labels: dict[str, str] = {}
    custom_order: list[str] = []
    for index, raw_result in enumerate(results, 1):
        if not isinstance(raw_result, dict):
            continue
        columns.append(str(raw_result.get("source_path") or raw_result.get("source_file") or f"结果 {index}"))
        metric_values: dict[str, typing.Any] = {"sample_count": raw_result.get("sample_count")}
        excluded = raw_result.get("excluded_counts") or {}
        metric_values.update({
            "excluded_above_limit": excluded.get("above_limit", 0),
            "excluded_negative": excluded.get("negative", 0),
            "excluded_invalid": excluded.get("invalid", 0),
        })
        for metric in raw_result.get("metrics") or []:
            if not isinstance(metric, dict) or not metric.get("key"):
                continue
            key = str(metric["key"])
            metric_values[key] = metric.get("value")
            labels[key] = str(metric.get("label") or key)
            if key not in METRIC_ORDER and key not in custom_order:
                custom_order.append(key)
        values_by_result.append(metric_values)
    if not columns:
        return None
    available = set().union(*(set(item) for item in values_by_result))
    keys = [key for key in METRIC_ORDER if key in available] + custom_order
    rows = [
        {
            "key": key,
            "label": METRIC_LABELS.get(key) or labels.get(key) or key,
            "values": [_format_number(item.get(key)) for item in values_by_result],
        }
        for key in keys
    ]
    script = summary.get("statistics_script") or {}
    return {
        "step_id": step.id,
        "step_name": step.name,
        "script_filename": script.get("filename") if isinstance(script, dict) else "",
        "max_latency_ns": summary.get("max_latency_ns"),
        "columns": columns,
        "rows": rows,
    }


def _comparison_number(value: typing.Any, unit: str = "", *, signed: bool = False) -> str:
    if value is None:
        return "-"
    try:
        number = float(value)
    except (TypeError, ValueError):
        return "-"
    sign = "+" if signed and number > 0 else ""
    rendered = f"{number:.3f}".rstrip("0").rstrip(".")
    return f"{sign}{rendered}{f' {unit}' if unit else ''}"


def _comparison_section(db: Session, run: TestRun) -> typing.Optional[dict[str, typing.Any]]:
    comparison = db.scalar(
        select(RunComparison).where(RunComparison.run_id == run.id)
    )
    if comparison is None:
        return None
    assessment_text = {
        "improved": "下降",
        "stable": "持平",
        "regressed": "上升",
        "added": "新增",
        "missing": "缺失",
        "incompatible": "不可比",
    }
    rows = []
    for raw in comparison.comparison_rows or []:
        if not isinstance(raw, dict):
            continue
        unit = str(raw.get("unit") or "")
        percentage = raw.get("percentage_delta")
        rows.append({
            **raw,
            "baseline_display": _comparison_number(raw.get("baseline_value"), unit),
            "target_display": _comparison_number(raw.get("target_value"), unit),
            "delta_display": _comparison_number(raw.get("absolute_delta"), unit, signed=True),
            "percentage_display": (
                _comparison_number(percentage, "%", signed=True)
                if percentage is not None else "基线为 0"
            ),
            "assessment_text": assessment_text.get(str(raw.get("assessment")), "未知"),
        })
    return {
        "target_run_number": comparison.target_run_number,
        "baseline_run_number": comparison.baseline_run_number,
        "compatibility": "统计配置一致" if comparison.is_compatible else "存在差异",
        "warnings": comparison.warnings or [],
        "saved_at": format_beijing(comparison.updated_at),
        "target_analysis_refs": comparison.target_analysis_refs or [],
        "baseline_analysis_refs": comparison.baseline_analysis_refs or [],
        "rows": rows,
    }


def build_report_document(
    db: Session, run: TestRun, report_version: int, generated_at: typing.Any,
    step: typing.Optional[RunStep] = None,
) -> dict[str, typing.Any]:
    cutoff = step.position if step is not None else None
    steps = [item for item in run.steps if cutoff is None or item.position < cutoff]
    servers, databases = _capture_sections(db, run, steps)
    orders = _order_sections(db, run, steps)
    statistics = [
        table for table in (_statistics_table(item) for item in steps if item.node_type == "data_statistics")
        if table is not None
    ]
    comparison = _comparison_section(db, run)
    snapshot = run.config_snapshot or {}
    plan = snapshot.get("plan") if isinstance(snapshot.get("plan"), dict) else {}
    scenario = snapshot.get("scenario") if isinstance(snapshot.get("scenario"), dict) else {}
    missing_sections = [
        label for label, values in (
            ("servers", servers), ("databases", databases), ("orders", orders), ("statistics", statistics)
        ) if not values
    ]
    verdict = run.verdict
    return {
        "report_version": report_version,
        "generated_at": format_beijing(generated_at),
        "run_number": run.run_number,
        "business_code": run.business_code,
        "status": run.status,
        "started_at": format_beijing(run.started_at) if run.started_at else "-",
        "finished_at": format_beijing(run.finished_at) if run.finished_at else "-",
        "plan": plan,
        "scenario": scenario,
        "servers": servers,
        "databases": databases,
        "orders": orders,
        "statistics": statistics,
        "comparison": comparison,
        "steps": [
            {
                "position": item.position,
                "name": item.name,
                "node_type": item.node_type,
                "status": item.status,
                "duration_ms": item.duration_ms if item.duration_ms is not None else "-",
                "error": item.error_message or "",
            }
            for item in run.steps
        ],
        "verdict": {
            "final_result": verdict.final_result if verdict and verdict.final_result else "待复核",
            "issue_description": verdict.issue_description if verdict else "",
            "notes": verdict.notes if verdict else "",
        },
        "missing_sections": missing_sections,
    }


def _render_html(report: dict[str, typing.Any], path: Path) -> None:
    environment = Environment(
        autoescape=select_autoescape(default=True),
        undefined=StrictUndefined,
        trim_blocks=True,
        lstrip_blocks=True,
    )
    path.write_text(environment.from_string(REPORT_TEMPLATE).render(report=report), encoding="utf-8")


def _style_sheet(sheet: typing.Any, widths: typing.Optional[list[int]] = None) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="277F74")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if widths:
        for index, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(index)].width = width


def _safe_sheet_title(workbook: Workbook, base: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", base)[:31] or "Sheet"
    candidate = cleaned
    counter = 2
    while candidate in workbook.sheetnames:
        suffix = f"-{counter}"
        candidate = cleaned[:31 - len(suffix)] + suffix
        counter += 1
    return candidate


def _render_xlsx(report: dict[str, typing.Any], path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "运行摘要"
    summary.append(["项目", "值"])
    summary_rows = (
        ("运行编号", report["run_number"]),
        ("报告版本", f"v{report['report_version']:03d}"),
        ("生成时间", report["generated_at"]),
        ("方案", report["plan"].get("name") or "-"),
        ("场景", report["scenario"].get("name") or "-"),
        ("业务", report["business_code"]),
        ("运行状态", report["status"]),
        ("开始时间", report["started_at"]),
        ("结束时间", report["finished_at"]),
        ("最终结论", report["verdict"]["final_result"]),
        ("问题说明", report["verdict"]["issue_description"]),
        ("备注", report["verdict"]["notes"]),
    )
    for row in summary_rows:
        summary.append(list(row))
    _style_sheet(summary, [22, 80])

    servers = workbook.create_sheet("服务器配置")
    servers.append(["步骤", "资源", "角色", "主机", "配置项", "采集值", "状态"])
    for source in report["servers"]:
        for item in source["items"]:
            servers.append([source["step_name"], source["resource_name"], source["resource_type"], source["host"], item["label"], item["value"], item["status"]])
    if not report["servers"]:
        servers.append(["无数据"])
    _style_sheet(servers, [22, 20, 14, 20, 22, 60, 12])

    databases = workbook.create_sheet("数据库配置")
    databases.append(["步骤", "数据库", "配置键", "采集值", "状态"])
    for source in report["databases"]:
        for item in source["items"]:
            databases.append([source["step_name"], source["database_name"], item["label"], item["value"], item["status"]])
    if not report["databases"]:
        databases.append(["无数据"])
    _style_sheet(databases, [22, 24, 42, 60, 12])

    orders = workbook.create_sheet("发单配置")
    orders.append(["步骤", "XML 文件", "SHA-256", "配置路径", "配置项", "值"])
    for order in report["orders"]:
        for row in order["rows"]:
            orders.append([order["step_name"], order["filename"], order["checksum"], row["path"], row["label"], row["value"]])
    if not report["orders"]:
        orders.append(["无数据"])
    _style_sheet(orders, [22, 32, 68, 52, 22, 44])

    for table_index, table in enumerate(report["statistics"], 1):
        sheet = workbook.create_sheet(_safe_sheet_title(workbook, f"指标-{table_index}-{table['step_name']}"))
        sheet.append(["异常上限(ns)", table["max_latency_ns"] if table["max_latency_ns"] is not None else "-"])
        sheet.append(["指标", *table["columns"]])
        for row in table["rows"]:
            sheet.append([row["label"], *row["values"]])
        _style_sheet(sheet, [18, *([32] * len(table["columns"]))])

    if report["comparison"]:
        comparison = workbook.create_sheet("运行对比")
        comparison.append(["当前运行", report["comparison"]["target_run_number"]])
        comparison.append(["基线运行", report["comparison"]["baseline_run_number"]])
        comparison.append(["快照保存时间", report["comparison"]["saved_at"]])
        comparison.append(["可比性", report["comparison"]["compatibility"]])
        comparison.append(["可比性说明", "；".join(report["comparison"]["warnings"]) or "-"])
        comparison.append([])
        comparison.append(["指标", "节点", "数据源", "基线值", "当前值", "绝对变化", "变化率", "判断"])
        for row in report["comparison"]["rows"]:
            comparison.append([
                row["metric_label"], row["step_name"], row["source_file"],
                row["baseline_display"], row["target_display"], row["delta_display"],
                row["percentage_display"], row["assessment_text"],
            ])
        _style_sheet(comparison, [22, 22, 34, 16, 16, 16, 14, 12])

    steps = workbook.create_sheet("步骤时间线")
    steps.append(["顺序", "步骤", "类型", "状态", "耗时(ms)", "错误"])
    for item in report["steps"]:
        steps.append([item["position"], item["name"], item["node_type"], item["status"], item["duration_ms"], item["error"]])
    _style_sheet(steps, [10, 24, 24, 16, 14, 56])

    for index, order in enumerate(report["orders"], 1):
        sheet = workbook.create_sheet(_safe_sheet_title(workbook, f"XML原文-{index}"))
        sheet.append([f"{order['step_name']} / {order['filename']}"])
        chunks: list[str] = []
        for line in order["raw_xml"].splitlines() or [""]:
            chunks.extend(line[offset:offset + 32000] for offset in range(0, max(1, len(line)), 32000))
        for chunk in chunks:
            sheet.append([chunk])
        _style_sheet(sheet, [120])
    workbook.save(path)


def _render_pdf(report: dict[str, typing.Any], path: Path) -> None:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import (
        BaseDocTemplate,
        Frame,
        KeepTogether,
        NextPageTemplate,
        PageBreak,
        PageTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
        XPreformatted,
    )

    font_name = "STSong-Light"
    try:
        pdfmetrics.getFont(font_name)
    except KeyError:
        pdfmetrics.registerFont(UnicodeCIDFont(font_name))
    styles = getSampleStyleSheet()
    body = ParagraphStyle("BodyCN", parent=styles["BodyText"], fontName=font_name, fontSize=8.2, leading=11)
    small = ParagraphStyle("SmallCN", parent=body, fontSize=7, leading=9, textColor=colors.HexColor("#52636B"))
    title = ParagraphStyle("TitleCN", parent=styles["Title"], fontName=font_name, fontSize=20, leading=25, textColor=colors.HexColor("#126B61"), spaceAfter=5)
    heading = ParagraphStyle("HeadingCN", parent=styles["Heading2"], fontName=font_name, fontSize=13, leading=17, textColor=colors.HexColor("#174D49"), spaceBefore=12, spaceAfter=6)
    subheading = ParagraphStyle("SubheadingCN", parent=styles["Heading3"], fontName=font_name, fontSize=10, leading=13, spaceBefore=8, spaceAfter=4)
    center = ParagraphStyle("CenterCN", parent=body, alignment=TA_CENTER)
    right = ParagraphStyle("RightCN", parent=body, alignment=TA_RIGHT)
    code = ParagraphStyle("CodeCN", parent=body, fontSize=6.2, leading=8)
    portrait_size = A4
    landscape_size = landscape(A4)
    margin_x, margin_top, margin_bottom = 14 * mm, 15 * mm, 15 * mm

    def footer(canvas: typing.Any, document: typing.Any) -> None:
        canvas.saveState()
        canvas.setFont(font_name, 7)
        canvas.setFillColor(colors.HexColor("#68777D"))
        canvas.drawString(margin_x, 8 * mm, report["run_number"])
        canvas.drawRightString(canvas._pagesize[0] - margin_x, 8 * mm, f"第 {document.page} 页")
        canvas.restoreState()

    document = BaseDocTemplate(
        str(path), pagesize=portrait_size, leftMargin=margin_x, rightMargin=margin_x,
        topMargin=margin_top, bottomMargin=margin_bottom,
        title=f"OpenSLT 测速报告 {report['run_number']}", author="OpenSLT",
    )
    portrait_frame = Frame(margin_x, margin_bottom, portrait_size[0] - 2 * margin_x, portrait_size[1] - margin_top - margin_bottom, id="portrait")
    landscape_frame = Frame(margin_x, margin_bottom, landscape_size[0] - 2 * margin_x, landscape_size[1] - margin_top - margin_bottom, id="landscape")
    document.addPageTemplates([
        PageTemplate(id="portrait", pagesize=portrait_size, frames=[portrait_frame], onPage=footer),
        PageTemplate(id="landscape", pagesize=landscape_size, frames=[landscape_frame], onPage=footer),
    ])

    def cell(value: typing.Any, style: ParagraphStyle = body) -> Paragraph:
        return Paragraph(html.escape(str(value if value not in (None, "") else "-")), style)

    def table(rows: list[list[typing.Any]], widths: typing.Optional[list[float]] = None, repeat: int = 1) -> Table:
        result = Table(rows, colWidths=widths, repeatRows=repeat, hAlign="LEFT")
        result.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("LEADING", (0, 0), (-1, -1), 10),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DCEBE8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#173B38")),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#87999F")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        return result

    story: list[typing.Any] = [
        Paragraph("OpenSLT 测速报告", title),
        Paragraph(f"运行编号：{html.escape(report['run_number'])}　报告版本：v{report['report_version']:03d}　生成时间：{html.escape(report['generated_at'])}", small),
        Spacer(1, 5 * mm),
        table([
            [cell("方案"), cell(report["plan"].get("name")), cell("场景"), cell(report["scenario"].get("name"))],
            [cell("业务"), cell(report["business_code"]), cell("运行状态"), cell(report["status"])],
            [cell("开始时间"), cell(report["started_at"]), cell("结束时间"), cell(report["finished_at"])],
        ], [25 * mm, 62 * mm, 25 * mm, 62 * mm], repeat=0),
    ]

    story.append(Paragraph("服务器配置", heading))
    if not report["servers"]:
        story.append(Paragraph("无服务器配置数据", body))
    for source in report["servers"]:
        story.extend([
            Paragraph(f"{html.escape(source['step_name'])} / {html.escape(source['resource_name'])}（{html.escape(source['resource_type'])}）", subheading),
            Paragraph(f"主机：{html.escape(source['host'] or '-')}　采集状态：{html.escape(source['status'])}", small),
            table([[cell("配置项"), cell("采集值"), cell("状态")]] + [[cell(item["label"]), cell(item["value"]), cell(item["status"])] for item in source["items"]], [38 * mm, 116 * mm, 20 * mm]),
        ])
    story.append(Paragraph("数据库配置", heading))
    if not report["databases"]:
        story.append(Paragraph("无数据库配置数据", body))
    for source in report["databases"]:
        story.extend([
            Paragraph(f"{html.escape(source['step_name'])} / {html.escape(source['database_name'] or '-')}", subheading),
            table([[cell("配置键"), cell("采集值"), cell("状态")]] + [[cell(item["label"]), cell(item["value"]), cell(item["status"])] for item in source["items"]], [58 * mm, 96 * mm, 20 * mm]),
        ])
    story.append(Paragraph("发单 XML 配置", heading))
    if not report["orders"]:
        story.append(Paragraph("无发单 XML 配置数据", body))
    for order in report["orders"]:
        story.extend([
            Paragraph(f"{html.escape(order['step_name'])} / {html.escape(order['filename'])}", subheading),
            Paragraph(f"SHA-256：{html.escape(order['checksum'])}", small),
            table([[cell("配置路径"), cell("配置项"), cell("值")]] + [[cell(row["path"], small), cell(row["label"]), cell(row["value"], small)] for row in order["rows"]], [72 * mm, 38 * mm, 64 * mm]),
        ])

    story.extend([NextPageTemplate("landscape"), PageBreak(), Paragraph("测速指标", heading)])
    if not report["statistics"]:
        story.append(Paragraph("无测速统计数据", body))
    for table_data in report["statistics"]:
        for start in range(0, len(table_data["columns"]), 6):
            columns = table_data["columns"][start:start + 6]
            rows = [[cell("指标", center), *[cell(value, center) for value in columns]]]
            for metric_row in table_data["rows"]:
                values = metric_row["values"][start:start + 6]
                rows.append([cell(metric_row["label"]), *[cell(value, right) for value in values]])
            usable_width = landscape_size[0] - 2 * margin_x
            metric_width = 27 * mm
            value_width = (usable_width - metric_width) / max(1, len(columns))
            block: list[typing.Any] = []
            if start == 0:
                block.extend([
                    Paragraph(html.escape(table_data["step_name"]), subheading),
                    Paragraph(
                        f"统计脚本：{html.escape(table_data['script_filename'] or '-')}　单位：ns　"
                        f"异常上限：{html.escape(str(table_data['max_latency_ns'] if table_data['max_latency_ns'] is not None else '-'))} ns",
                        small,
                    ),
                ])
            block.extend([
                table(rows, [metric_width, *([value_width] * len(columns))]),
                Spacer(1, 4 * mm),
            ])
            story.append(KeepTogether(block))

    if report["comparison"]:
        comparison = report["comparison"]
        story.extend([
            Paragraph("运行对比", heading),
            Paragraph(
                f"当前运行：{html.escape(comparison['target_run_number'])}　"
                f"基线运行：{html.escape(comparison['baseline_run_number'])}　"
                f"快照保存时间：{html.escape(comparison['saved_at'])}　"
                f"可比性：{html.escape(comparison['compatibility'])}",
                small,
            ),
        ])
        if comparison["warnings"]:
            story.append(Paragraph(
                f"可比性说明：{html.escape('；'.join(comparison['warnings']))}", small
            ))
        comparison_rows = [[
            cell("指标"), cell("节点 / 数据源"), cell("基线值"), cell("当前值"),
            cell("绝对变化"), cell("变化率"), cell("判断"),
        ]]
        comparison_rows.extend([
            cell(row["metric_label"]),
            cell(f"{row['step_name']} / {row['source_file']}", small),
            cell(row["baseline_display"], right),
            cell(row["target_display"], right),
            cell(row["delta_display"], right),
            cell(row["percentage_display"], right),
            cell(row["assessment_text"]),
        ] for row in comparison["rows"])
        story.append(table(
            comparison_rows,
            [33 * mm, 62 * mm, 28 * mm, 28 * mm, 28 * mm, 24 * mm, 22 * mm],
        ))

    story.extend([NextPageTemplate("portrait"), PageBreak(), Paragraph("步骤时间线", heading)])
    story.append(table(
        [[cell("顺序"), cell("步骤"), cell("类型"), cell("状态"), cell("耗时(ms)"), cell("错误")]]
        + [[cell(item["position"]), cell(item["name"]), cell(item["node_type"]), cell(item["status"]), cell(item["duration_ms"], right), cell(item["error"])] for item in report["steps"]],
        [12 * mm, 34 * mm, 34 * mm, 24 * mm, 22 * mm, 48 * mm],
    ))
    story.append(Paragraph("最终结论", heading))
    story.append(table([
        [cell("结论"), cell(report["verdict"]["final_result"])],
        [cell("问题说明"), cell(report["verdict"]["issue_description"])],
        [cell("备注"), cell(report["verdict"]["notes"])],
    ], [32 * mm, 142 * mm], repeat=0))
    if report["orders"]:
        story.append(PageBreak())
        story.append(Paragraph("发单 XML 原文附录", heading))
        for order in report["orders"]:
            story.append(Paragraph(f"{html.escape(order['step_name'])} / {html.escape(order['filename'])}", subheading))
            wrapped_lines: list[str] = []
            for raw_line in order["raw_xml"].splitlines() or [""]:
                wrapped_lines.extend(
                    raw_line[offset:offset + 75]
                    for offset in range(0, max(1, len(raw_line)), 75)
                )
            story.append(XPreformatted(html.escape("\n".join(wrapped_lines)), code))
    document.build(story)


def _next_report_version(db: Session, run: TestRun, root: Path) -> int:
    versions = [
        int(match.group(1))
        for name in db.scalars(select(Artifact.name).where(
            Artifact.run_id == run.id,
            Artifact.artifact_type.in_(REPORT_ARTIFACT_TYPES),
        )).all()
        for match in [REPORT_VERSION_RE.fullmatch(name)]
        if match
    ]
    if root.is_dir():
        versions.extend(
            int(path.name[1:])
            for path in root.iterdir()
            if path.is_dir() and re.fullmatch(r"v\d+", path.name)
        )
    return max(versions, default=0) + 1


def _register_report_artifact(
    db: Session, run: TestRun, step: typing.Optional[RunStep], path: Path,
    artifact_type: str, content_type: str,
) -> Artifact:
    data = path.read_bytes()
    artifact = Artifact(
        run_id=run.id,
        step_id=step.id if step else None,
        artifact_type=artifact_type,
        name=path.name,
        path=str(path),
        content_type=content_type,
        size=len(data),
        checksum=hashlib.sha256(data).hexdigest(),
        is_immutable=True,
    )
    db.add(artifact)
    db.flush()
    return artifact


def generate_reports(
    db: Session,
    run: TestRun,
    *,
    step: typing.Optional[RunStep] = None,
    reason: str = "manual",
) -> dict[str, typing.Any]:
    db.scalar(select(TestRun.id).where(TestRun.id == run.id).with_for_update())
    root = (
        settings.artifact_root / run.business_code / str(run.plan_id) / str(run.scenario_id)
        / run.run_number / "reports"
    )
    root.mkdir(parents=True, exist_ok=True)
    report_version = _next_report_version(db, run, root)
    generated_at = beijing_now()
    report = build_report_document(db, run, report_version, generated_at, step)
    directory = root / f"v{report_version:03d}"
    staging = root / f".v{report_version:03d}.{uuid4().hex}.tmp"
    filenames = {
        "web_report": f"report-v{report_version:03d}.html",
        "excel_report": f"report-v{report_version:03d}.xlsx",
        "pdf_report": f"report-v{report_version:03d}.pdf",
    }
    try:
        staging.mkdir(parents=False, exist_ok=False)
        _render_html(report, staging / filenames["web_report"])
        _render_xlsx(report, staging / filenames["excel_report"])
        _render_pdf(report, staging / filenames["pdf_report"])
        if directory.exists():
            raise RuntimeError(f"报告版本目录已存在: {directory}")
        staging.replace(directory)
    except Exception:
        shutil.rmtree(staging, ignore_errors=True)
        raise
    artifacts = [
        _register_report_artifact(db, run, step, directory / filenames["web_report"], "web_report", "text/html; charset=utf-8"),
        _register_report_artifact(db, run, step, directory / filenames["excel_report"], "excel_report", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        _register_report_artifact(db, run, step, directory / filenames["pdf_report"], "pdf_report", "application/pdf"),
    ]
    result = {
        "report_version": report_version,
        "generated_at": generated_at.isoformat(),
        "reason": reason,
        "artifact_ids": [artifact.id for artifact in artifacts],
        "artifacts": [
            {"id": artifact.id, "name": artifact.name, "artifact_type": artifact.artifact_type, "checksum": artifact.checksum}
            for artifact in artifacts
        ],
        "missing_sections": report["missing_sections"],
    }
    if step is not None:
        step.result_summary = result
    db.flush()
    return result
