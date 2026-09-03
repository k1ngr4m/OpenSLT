from __future__ import annotations

import errno
import hashlib
import json
import os
import pty
import re
import select as io_select
import shutil
import sqlite3
import struct
import subprocess
import termios
import time
import typing
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime
from pathlib import Path, PurePosixPath
from urllib.parse import quote, urlsplit, urlunsplit
from uuid import uuid4

from sqlalchemy import select
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from pypdf import PdfReader

from app.core.config import settings
from app.core.database import SessionLocal
from app.core.logging import redact
from app.core.security import decrypt_secret
from app.core.time import beijing_now
from app.models import DurableTask, SvnKnowledgeSource
from app.services.durable_tasks import enqueue_task
from app.services.embedding import EmbeddingClient


SUPPORTED_SUFFIXES = {
    ".csv", ".docx", ".htm", ".html", ".json", ".md", ".pdf", ".txt",
    ".xlsx", ".yaml", ".yml",
}
ACTIVE_TASK_STATUSES = {"queued", "running"}


class SvnKnowledgeError(RuntimeError):
    pass


def normalize_repository_url(value: str, allow_insecure_http: bool) -> str:
    raw = value.strip().rstrip("/")
    parts = urlsplit(raw)
    if parts.scheme not in {"http", "https"} or not parts.netloc:
        raise ValueError("SVN 仓库地址必须是完整的 HTTP 或 HTTPS URL")
    if parts.username or parts.password:
        raise ValueError("SVN 仓库地址不得包含账号或密码")
    if parts.query or parts.fragment:
        raise ValueError("SVN 仓库地址不得包含查询参数或片段")
    if parts.scheme == "http" and not allow_insecure_http:
        raise ValueError("HTTP 会明文传输凭据，必须由管理员显式允许后才能保存")
    return urlunsplit((parts.scheme, parts.netloc, parts.path.rstrip("/"), "", ""))


def normalize_repository_urls(values: typing.Iterable[str], allow_insecure_http: bool) -> typing.List[str]:
    result = list(dict.fromkeys(normalize_repository_url(value, allow_insecure_http) for value in values))
    if not result:
        raise ValueError("至少配置一个 SVN 仓库 URL")
    return result


def normalize_include_paths(values: typing.Iterable[str]) -> typing.List[str]:
    result: typing.List[str] = []
    seen = set()
    for value in values:
        parts = urlsplit(value.strip())
        if parts.scheme or parts.netloc:
            raise ValueError("白名单路径必须是 SVN 相对路径")
        raw = value.strip().replace("\\", "/").strip("/")
        path = PurePosixPath(raw)
        if not raw or value.strip().startswith(("/", "\\")) or any(part in {"", ".", "..", ".svn"} for part in path.parts):
            raise ValueError("白名单路径必须是非空 SVN 相对路径，且不得包含 .、.. 或 .svn")
        normalized = path.as_posix()
        if normalized not in seen:
            seen.add(normalized)
            result.append(normalized)
    if not result:
        raise ValueError("至少配置一个允许索引的 SVN 相对路径")
    for parent in result:
        if any(child != parent and child.startswith(parent + "/") for child in result):
            raise ValueError("白名单路径不得互相包含，请只保留更精确的子目录")
    return result


def join_repository_url(repository_url: str, relative_path: str) -> str:
    return repository_url.rstrip("/") + "/" + quote(relative_path, safe="/")


def _svn_targets(
    repository_urls: typing.Sequence[str],
    include_paths: typing.Sequence[str],
) -> typing.List[typing.Tuple[str, str, str]]:
    targets: typing.List[typing.Tuple[str, str, str]] = []
    seen = set()
    for include_path in include_paths:
        for repository_url in repository_urls:
            target_url = join_repository_url(repository_url, include_path)
            if target_url in seen:
                continue
            seen.add(target_url)
            source_ref = include_path if len(repository_urls) == 1 else target_url
            targets.append((repository_url, include_path, source_ref))
    return targets


class SvnClient:
    """Run SVN without putting a password in argv, env, logs, or auth cache."""

    def __init__(self, executable: str = "svn", timeout_seconds: int = 120) -> None:
        self.executable = executable
        self.timeout_seconds = timeout_seconds

    def version(self) -> str:
        executable = shutil.which(self.executable)
        if not executable:
            raise SvnKnowledgeError("未安装 SVN 客户端，请在离线包中安装 subversion RPM")
        completed = subprocess.run(
            [executable, "--version", "--quiet"],
            capture_output=True,
            text=True,
            timeout=10,
            check=False,
        )
        if completed.returncode:
            raise SvnKnowledgeError("SVN 客户端就绪检查失败")
        return completed.stdout.strip()

    def run(self, arguments: typing.Sequence[str], username: str, password: str) -> str:
        executable = shutil.which(self.executable)
        if not executable:
            raise SvnKnowledgeError("未安装 SVN 客户端，请在离线包中安装 subversion RPM")
        if not arguments:
            raise ValueError("SVN 子命令不能为空")
        config_directory = settings.knowledge_root / "svn-client-config"
        config_directory.mkdir(parents=True, exist_ok=True)
        try:
            config_directory.chmod(0o700)
        except OSError:
            pass
        command = [
            executable,
            arguments[0],
            "--username", username,
            "--config-dir", str(config_directory),
            "--no-auth-cache",
            "--config-option", "config:auth:store-auth-creds=no",
            "--config-option", "config:auth:store-passwords=no",
            *arguments[1:],
        ]
        master, slave = pty.openpty()
        attributes = termios.tcgetattr(slave)
        attributes[3] &= ~termios.ECHO
        termios.tcsetattr(slave, termios.TCSANOW, attributes)
        environment = dict(os.environ)
        environment.update({"LC_ALL": "C", "LANG": "C"})
        process = subprocess.Popen(
            command,
            stdin=slave,
            stdout=slave,
            stderr=slave,
            env=environment,
            close_fds=True,
        )
        os.close(slave)
        output = bytearray()
        prompted = False
        deadline = time.monotonic() + self.timeout_seconds
        try:
            while True:
                if time.monotonic() >= deadline:
                    process.terminate()
                    raise SvnKnowledgeError("SVN 请求超时")
                ready, _, _ = io_select.select([master], [], [], 0.1)
                if ready:
                    try:
                        chunk = os.read(master, 4096)
                    except OSError as exc:
                        if exc.errno != errno.EIO:
                            raise
                        chunk = b""
                    output.extend(chunk)
                    lowered = bytes(output[-1024:]).lower()
                    if not prompted and b"password for" in lowered:
                        os.write(master, password.encode("utf-8") + b"\n")
                        prompted = True
                if process.poll() is not None:
                    while True:
                        try:
                            chunk = os.read(master, 4096)
                        except OSError as exc:
                            if exc.errno == errno.EIO:
                                break
                            raise
                        if not chunk:
                            break
                        output.extend(chunk)
                    break
        finally:
            os.close(master)
            if process.poll() is None:
                process.kill()
            process.wait()
        text = output.decode("utf-8", errors="replace")
        if process.returncode:
            lowered = text.casefold()
            if "authentication failed" in lowered or "authorization failed" in lowered or "could not authenticate" in lowered:
                raise SvnKnowledgeError("SVN 认证失败，请检查只读账号、密码和目录权限")
            if "conflict" in lowered:
                raise SvnKnowledgeError("SVN working copy 存在冲突，已停止同步且未覆盖本地文件")
            raise SvnKnowledgeError("SVN 命令执行失败（退出码 %s）" % process.returncode)
        return text


def svn_client_status(client: typing.Optional[SvnClient] = None) -> typing.Dict[str, typing.Any]:
    try:
        return {"ready": True, "version": (client or SvnClient()).version(), "error": None}
    except (OSError, subprocess.SubprocessError, SvnKnowledgeError) as exc:
        return {"ready": False, "version": None, "error": redact(str(exc))}


def _xml_root(output: str) -> ET.Element:
    start = output.find("<?xml")
    if start < 0:
        raise SvnKnowledgeError("SVN 返回了无法识别的响应")
    try:
        return ET.fromstring(output[start:])
    except ET.ParseError as exc:
        raise SvnKnowledgeError("SVN 返回了无法解析的 XML 响应") from exc


def test_svn_connection(
    repository_urls: typing.Sequence[str],
    username: str,
    password: str,
    include_paths: typing.Sequence[str],
    client: typing.Optional[SvnClient] = None,
) -> typing.Dict[str, typing.Any]:
    runner = client or SvnClient()
    version = runner.version()
    for repository_url in repository_urls:
        _xml_root(runner.run(["info", "--xml", repository_url], username, password))
    targets = _svn_targets(repository_urls, include_paths)
    for repository_url, relative_path, _ in targets:
        _xml_root(runner.run(["list", "--xml", join_repository_url(repository_url, relative_path)], username, password))
    return {"ok": True, "svn_version": version, "checked_paths": [item[2] for item in targets]}


def _working_copy_path(repository_url: str, relative_path: str) -> Path:
    key = hashlib.sha256((repository_url + "\n" + relative_path).encode("utf-8")).hexdigest()[:20]
    return settings.knowledge_root / "svn-working-copies" / key


def _working_copy_info(client: SvnClient, path: Path, username: str, password: str) -> typing.Tuple[str, str]:
    root = _xml_root(client.run(["info", "--xml", str(path)], username, password))
    entry = root.find("entry")
    if entry is None or entry.findtext("url") is None:
        raise SvnKnowledgeError("SVN working copy 信息不完整")
    return entry.findtext("url", "").rstrip("/"), entry.attrib.get("revision", "")


def _sync_working_copy(
    client: SvnClient,
    repository_url: str,
    relative_path: str,
    username: str,
    password: str,
) -> typing.Tuple[Path, str]:
    target_url = join_repository_url(repository_url, relative_path)
    working_copy = _working_copy_path(repository_url, relative_path)
    working_copy.parent.mkdir(parents=True, exist_ok=True)
    if working_copy.exists():
        if not (working_copy / ".svn").is_dir():
            raise SvnKnowledgeError("SVN working copy 目录已存在但不是有效检出目录")
        actual_url, _ = _working_copy_info(client, working_copy, username, password)
        if actual_url != target_url.rstrip("/"):
            raise SvnKnowledgeError("SVN working copy URL 与当前配置不一致，已停止同步")
        status = _xml_root(client.run(["status", "--xml", str(working_copy)], username, password))
        if status.findall(".//entry"):
            raise SvnKnowledgeError("SVN working copy 存在本地修改或冲突，已停止同步且未覆盖本地文件")
        client.run(["update", "--ignore-externals", str(working_copy)], username, password)
    else:
        temporary = working_copy.with_name(working_copy.name + "." + uuid4().hex + ".checkout")
        try:
            client.run(["checkout", "--ignore-externals", target_url, str(temporary)], username, password)
            os.replace(str(temporary), str(working_copy))
        finally:
            if temporary.exists():
                shutil.rmtree(temporary)
    actual_url, revision = _working_copy_info(client, working_copy, username, password)
    if actual_url != target_url.rstrip("/"):
        raise SvnKnowledgeError("SVN checkout URL 校验失败")
    return working_copy, revision


def _load_manifest(path: Path) -> typing.Dict[str, typing.Any]:
    try:
        connection = sqlite3.connect(str(path))
        try:
            row = connection.execute("SELECT value FROM metadata WHERE key = 'manifest'").fetchone()
        finally:
            connection.close()
        value = json.loads(row[0]) if row else {}
        return value if isinstance(value, dict) else {}
    except (FileNotFoundError, json.JSONDecodeError, OSError, sqlite3.Error):
        return {}


def _manifest_repository_urls(manifest: typing.Dict[str, typing.Any]) -> typing.List[str]:
    repository_urls = manifest.get("repository_urls")
    if isinstance(repository_urls, list):
        return [str(value) for value in repository_urls]
    repository_url = manifest.get("repository_url")
    return [str(repository_url)] if repository_url else []


def published_index_matches(
    source: SvnKnowledgeSource,
    embedding_base_url: str,
    embedding_model: str,
) -> bool:
    path = settings.knowledge_root / "published" / "svn-index.sqlite3"
    manifest = _load_manifest(path)
    if not manifest:
        return False
    try:
        connection = sqlite3.connect(str(path))
        try:
            metadata = dict(connection.execute("SELECT key, value FROM metadata WHERE key IN ('embedding_model', 'embedding_base_url')"))
        finally:
            connection.close()
    except sqlite3.Error:
        return False
    repository_urls = list(source.repository_urls or []) or [source.repository_url]
    return (
        _manifest_repository_urls(manifest) == repository_urls
        and set(manifest.get("revisions", {})) == {item[2] for item in _svn_targets(repository_urls, source.include_paths)}
        and metadata.get("embedding_model") == embedding_model
        and metadata.get("embedding_base_url") == embedding_base_url
    )


def _build_manifest(
    repository_urls: typing.Union[str, typing.Sequence[str]],
    revisions: typing.Dict[str, str],
    working_copies: typing.Dict[str, Path],
    previous: typing.Dict[str, typing.Any],
) -> typing.Tuple[typing.Dict[str, typing.Any], typing.Dict[str, int]]:
    repositories = [repository_urls] if isinstance(repository_urls, str) else list(repository_urls)
    previous_files = previous.get("files") if _manifest_repository_urls(previous) == repositories else {}
    previous_files = previous_files if isinstance(previous_files, dict) else {}
    files: typing.Dict[str, typing.Any] = {}
    changes = {"added": 0, "changed": 0, "deleted": 0, "unchanged": 0}
    for include_path, root in working_copies.items():
        for directory, names, filenames in os.walk(root):
            names[:] = [name for name in names if name != ".svn"]
            for filename in filenames:
                source = Path(directory) / filename
                if source.is_symlink() or source.suffix.casefold() not in SUPPORTED_SUFFIXES:
                    continue
                relative = source.relative_to(root).as_posix()
                source_ref = include_path.rstrip("/") + "/" + relative
                stat = source.stat()
                old = previous_files.get(source_ref, {})
                item = {"size": stat.st_size, "mtime_ns": stat.st_mtime_ns}
                if old.get("size") == item["size"] and old.get("mtime_ns") == item["mtime_ns"] and old.get("sha256"):
                    item["sha256"] = old["sha256"]
                    changes["unchanged"] += 1
                else:
                    digest = hashlib.sha256()
                    with source.open("rb") as handle:
                        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                            digest.update(chunk)
                    item["sha256"] = digest.hexdigest()
                    changes["changed" if source_ref in previous_files else "added"] += 1
                files[source_ref] = item
    changes["deleted"] = len(set(previous_files) - set(files))
    return {
        "repository_url": repositories[0],
        "repository_urls": repositories,
        "revisions": revisions,
        "published_at": beijing_now().isoformat(),
        "files": files,
    }, changes


def _extract_text(path: Path) -> str:
    if path.stat().st_size > 50 * 1024 * 1024:
        raise ValueError("文件超过 50 MiB 索引限制")
    suffix = path.suffix.casefold()
    if suffix in {".txt", ".md", ".csv", ".json", ".yaml", ".yml", ".htm", ".html"}:
        raw = path.read_bytes()
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                return raw.decode(encoding)
            except UnicodeDecodeError:
                continue
        return raw.decode("utf-8", errors="replace")
    if suffix == ".docx":
        with zipfile.ZipFile(path) as archive:
            if sum(item.file_size for item in archive.infolist()) > 200 * 1024 * 1024:
                raise ValueError("DOCX 解压内容超过 200 MiB 限制")
            root = ET.fromstring(archive.read("word/document.xml"))
        return "\n".join(text.strip() for text in root.itertext() if text.strip())
    if suffix == ".xlsx":
        with zipfile.ZipFile(path) as archive:
            if sum(item.file_size for item in archive.infolist()) > 200 * 1024 * 1024:
                raise ValueError("XLSX 解压内容超过 200 MiB 限制")
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            lines = []
            for sheet in workbook.worksheets:
                lines.append("工作表：" + sheet.title)
                for row in sheet.iter_rows(values_only=True):
                    values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                    if values:
                        lines.append("\t".join(values))
            return "\n".join(lines)
        finally:
            workbook.close()
    if suffix == ".pdf":
        return "\n".join((page.extract_text() or "") for page in PdfReader(str(path)).pages)
    raise ValueError("不支持的知识文件格式")


def _chunks(text: str, size: int = 1200, overlap: int = 150) -> typing.Iterator[str]:
    normalized = "\n".join(line.strip() for line in text.splitlines() if line.strip())
    if len(normalized) > 5_000_000:
        raise ValueError("文件提取文本超过 500 万字符限制")
    start = 0
    while start < len(normalized):
        chunk = normalized[start:start + size].strip()
        if chunk:
            yield chunk
        start += size - overlap


def _local_source_path(source_ref: str, working_copies: typing.Dict[str, Path]) -> Path:
    include_path = max((path for path in working_copies if source_ref == path or source_ref.startswith(path + "/")), key=len)
    relative = source_ref[len(include_path):].lstrip("/")
    return working_copies[include_path] / relative


def _revision_for_source(source_ref: str, revisions: typing.Dict[str, str]) -> str:
    include_path = max((path for path in revisions if source_ref == path or source_ref.startswith(path + "/")), key=len)
    return revisions[include_path]


def _publish_vector_index(
    manifest: typing.Dict[str, typing.Any],
    previous: typing.Dict[str, typing.Any],
    working_copies: typing.Dict[str, Path],
    embedding: EmbeddingClient,
) -> int:
    destination = settings.knowledge_root / "published" / "svn-index.sqlite3"
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_name(destination.name + "." + uuid4().hex + ".tmp")
    if destination.exists():
        shutil.copy2(destination, temporary)
    try:
        connection = sqlite3.connect(str(temporary))
        try:
            connection.executescript(
                "CREATE TABLE IF NOT EXISTS metadata (key TEXT PRIMARY KEY, value TEXT NOT NULL);"
                "CREATE TABLE IF NOT EXISTS files (source_path TEXT PRIMARY KEY, size INTEGER NOT NULL, mtime_ns INTEGER NOT NULL, sha256 TEXT NOT NULL, revision TEXT NOT NULL);"
                "CREATE TABLE IF NOT EXISTS chunks (source_path TEXT NOT NULL, chunk_no INTEGER NOT NULL, content TEXT NOT NULL, vector BLOB NOT NULL, dimensions INTEGER NOT NULL, PRIMARY KEY(source_path, chunk_no));"
            )
            old_model = connection.execute("SELECT value FROM metadata WHERE key = 'embedding_model'").fetchone()
            old_base_url = connection.execute("SELECT value FROM metadata WHERE key = 'embedding_base_url'").fetchone()
            model_changed = (
                not old_model
                or old_model[0] != embedding.model
                or not old_base_url
                or old_base_url[0] != embedding.base_url
                or _manifest_repository_urls(previous) != _manifest_repository_urls(manifest)
            )
            previous_files = previous.get("files", {}) if not model_changed else {}
            current_files = manifest["files"]
            changed = {
                path for path, item in current_files.items()
                if path not in previous_files or previous_files[path].get("sha256") != item["sha256"] or path in previous.get("failed_files", {})
            }
            removed = set(previous_files) - set(current_files)
            if model_changed:
                connection.execute("DELETE FROM chunks")
                connection.execute("DELETE FROM files")
            else:
                for source_path in changed | removed:
                    connection.execute("DELETE FROM chunks WHERE source_path = ?", (source_path,))
                    connection.execute("DELETE FROM files WHERE source_path = ?", (source_path,))

            dimensions = 0
            failed_files: typing.Dict[str, str] = {}
            for source_path in sorted(changed):
                try:
                    texts = list(_chunks(_extract_text(_local_source_path(source_path, working_copies))))
                except Exception as exc:
                    failed_files[source_path] = str(redact(str(exc)))[:300]
                    continue
                for offset in range(0, len(texts), 16):
                    batch = texts[offset:offset + 16]
                    vectors = embedding.embed(batch)
                    dimensions = len(vectors[0])
                    for index, (content, vector) in enumerate(zip(batch, vectors), start=offset):
                        blob = struct.pack("<%sf" % len(vector), *vector)
                        connection.execute(
                            "INSERT INTO chunks(source_path, chunk_no, content, vector, dimensions) VALUES (?, ?, ?, ?, ?)",
                            (source_path, index, content, blob, len(vector)),
                        )
                item = current_files[source_path]
                connection.execute(
                    "INSERT INTO files(source_path, size, mtime_ns, sha256, revision) VALUES (?, ?, ?, ?, ?)",
                    (source_path, item["size"], item["mtime_ns"], item["sha256"], _revision_for_source(source_path, manifest["revisions"])),
                )
            for source_path in sorted(set(current_files) - changed):
                item = current_files[source_path]
                connection.execute(
                    "INSERT OR REPLACE INTO files(source_path, size, mtime_ns, sha256, revision) VALUES (?, ?, ?, ?, ?)",
                    (source_path, item["size"], item["mtime_ns"], item["sha256"], _revision_for_source(source_path, manifest["revisions"])),
                )
            manifest["failed_files"] = failed_files
            if not dimensions:
                row = connection.execute("SELECT dimensions FROM chunks LIMIT 1").fetchone()
                dimensions = int(row[0]) if row else 0
            if connection.execute("SELECT 1 FROM chunks LIMIT 1").fetchone() is None:
                raise SvnKnowledgeError("没有成功提取并向量化的知识内容，未发布空索引")
            connection.execute("INSERT OR REPLACE INTO metadata(key, value) VALUES ('manifest', ?)", (json.dumps(manifest, ensure_ascii=False, sort_keys=True),))
            connection.execute("INSERT OR REPLACE INTO metadata(key, value) VALUES ('embedding_model', ?)", (embedding.model,))
            connection.execute("INSERT OR REPLACE INTO metadata(key, value) VALUES ('embedding_base_url', ?)", (embedding.base_url,))
            connection.commit()
        finally:
            connection.close()
        os.replace(str(temporary), str(destination))
        try:
            destination.chmod(0o600)
        except OSError:
            pass
        return dimensions
    finally:
        try:
            temporary.unlink()
        except FileNotFoundError:
            pass


def search_vector_index(query: str, query_vector: typing.Sequence[float], top_k: int) -> typing.List[typing.Dict[str, typing.Any]]:
    index_path = settings.knowledge_root / "published" / "svn-index.sqlite3"
    if not index_path.exists():
        raise SvnKnowledgeError("尚无成功发布的知识索引")
    norm = sum(value * value for value in query_vector) ** 0.5
    if not norm:
        raise SvnKnowledgeError("Embedding 查询向量无效")
    terms = [term for term in query.casefold().split() if term]
    results = []
    connection = sqlite3.connect(str(index_path))
    try:
        stored_dimension = connection.execute("SELECT dimensions FROM chunks LIMIT 1").fetchone()
        if stored_dimension and int(stored_dimension[0]) != len(query_vector):
            raise SvnKnowledgeError("查询 embedding 维度与已发布索引不一致，请重新同步")
        rows = connection.execute(
            "SELECT c.source_path, f.revision, c.content, c.vector, c.dimensions FROM chunks c JOIN files f ON f.source_path = c.source_path"
        )
        # ponytail: application-layer scan is enough for the current corpus; add a vector DB only after measured latency exceeds the requirement.
        for source_path, revision, content, blob, dimensions in rows:
            if dimensions != len(query_vector):
                continue
            vector = struct.unpack("<%sf" % dimensions, blob)
            vector_norm = sum(value * value for value in vector) ** 0.5
            vector_score = sum(left * right for left, right in zip(query_vector, vector)) / (norm * vector_norm) if vector_norm else 0.0
            lowered = content.casefold()
            keyword_score = min(1.0, sum(lowered.count(term) for term in terms) / max(1, len(terms))) if terms else 0.0
            score = vector_score * 0.85 + keyword_score * 0.15
            results.append({
                "source_path": source_path,
                "revision": revision,
                "snippet": content[:500],
                "score": round(score, 6),
                "vector_score": round(vector_score, 6),
                "keyword_score": round(keyword_score, 6),
            })
    finally:
        connection.close()
    return sorted(results, key=lambda item: item["score"], reverse=True)[:top_k]


def list_indexed_requirements(query: str = "") -> typing.List[typing.Dict[str, typing.Any]]:
    index_path = settings.knowledge_root / "published" / "svn-index.sqlite3"
    if not index_path.exists():
        raise SvnKnowledgeError("尚无成功发布的知识索引")
    needle = query.strip().casefold()
    connection = sqlite3.connect(str(index_path))
    try:
        rows = connection.execute("SELECT source_path, revision FROM files ORDER BY source_path").fetchall()
    finally:
        connection.close()
    result = []
    for source_path, revision in rows:
        stem = PurePosixPath(source_path).stem
        match = re.search(r"(?<![A-Za-z0-9])([A-Za-z]{0,8}[-_]?\d{2,})(?!\d)", stem)
        requirement_no = match.group(1) if match else None
        if not requirement_no and not any(marker in source_path.casefold() for marker in ("需求", "requirement", "prd")):
            continue
        name = re.sub(r"^[A-Za-z]{0,8}[-_]?\d{2,}[\s._-]*", "", stem).strip() or stem
        name = re.sub(r"[\s_-]*(需求规格说明书|需求说明书|需求文档|需求)$", "", name).strip() or stem
        if needle and needle not in source_path.casefold() and needle not in name.casefold() and needle not in (requirement_no or "").casefold():
            continue
        result.append({"source_path": source_path, "revision": revision, "requirement_no": requirement_no, "requirement_name": name[:255]})
    return result[:500]


def get_indexed_document(source_path: str) -> typing.Dict[str, str]:
    index_path = settings.knowledge_root / "published" / "svn-index.sqlite3"
    connection = sqlite3.connect(str(index_path))
    try:
        row = connection.execute("SELECT revision FROM files WHERE source_path = ?", (source_path,)).fetchone()
        if row is None:
            raise SvnKnowledgeError("选中的需求不在当前知识索引中")
        chunks = connection.execute("SELECT content FROM chunks WHERE source_path = ? ORDER BY chunk_no", (source_path,)).fetchall()
    finally:
        connection.close()
    if not chunks:
        raise SvnKnowledgeError("选中的需求没有可用正文")
    return {"source_path": source_path, "revision": row[0], "content": "\n".join(item[0] for item in chunks)}


def active_svn_task(db: Session) -> typing.Optional[DurableTask]:
    return db.scalar(
        select(DurableTask)
        .where(DurableTask.task_type == "svn_sync", DurableTask.status.in_(ACTIVE_TASK_STATUSES))
        .order_by(DurableTask.id.desc())
    )


def enqueue_svn_sync(db: Session, source: SvnKnowledgeSource, reason: str, now: typing.Optional[datetime] = None) -> DurableTask:
    from app.services.model_providers import require_active_model

    require_active_model(db, "embedding")
    locked = db.scalar(select(SvnKnowledgeSource).where(SvnKnowledgeSource.id == source.id).with_for_update())
    if locked is None:
        raise SvnKnowledgeError("SVN 知识源尚未配置")
    existing = active_svn_task(db)
    if existing:
        return existing
    timestamp = now or beijing_now()
    if reason == "scheduled":
        window = int(timestamp.timestamp()) // (locked.sync_interval_minutes * 60)
        key = "svn-sync:%s:scheduled:%s" % (locked.id, window)
    else:
        key = "svn-sync:%s:manual:%s" % (locked.id, uuid4().hex)
    task = enqueue_task(db, "svn_sync", {"source_id": locked.id, "reason": reason}, key)
    if task.status in ACTIVE_TASK_STATUSES:
        locked.sync_status = "queued"
    db.commit()
    return task


def enqueue_due_svn_syncs(db: Session, now: typing.Optional[datetime] = None) -> typing.Optional[DurableTask]:
    source = db.scalar(select(SvnKnowledgeSource).where(SvnKnowledgeSource.enabled.is_(True)))
    if source is None:
        return None
    return enqueue_svn_sync(db, source, "scheduled", now)


def execute_svn_sync(source_id: int, client: typing.Optional[SvnClient] = None) -> None:
    from app.services.model_providers import require_active_model

    db = SessionLocal()
    try:
        source = db.get(SvnKnowledgeSource, source_id)
        if source is None:
            raise SvnKnowledgeError("SVN 知识源不存在")
        repository_urls = list(source.repository_urls or []) or [source.repository_url]
        username = source.username
        password = decrypt_secret(source.encrypted_password) or ""
        embedding_provider, embedding = require_active_model(db, "embedding")
        embedding_base_url = embedding_provider.base_url
        embedding_model = embedding.model_id
        embedding_api_key = decrypt_secret(embedding_provider.encrypted_api_key)
        include_paths = list(source.include_paths)
        source.sync_status = "running"
        source.last_attempt_at = beijing_now()
        db.commit()
    finally:
        db.close()

    try:
        runner = client or SvnClient()
        runner.version()
        working_copies: typing.Dict[str, Path] = {}
        revisions: typing.Dict[str, str] = {}
        for repository_url, relative_path, source_ref in _svn_targets(repository_urls, include_paths):
            working_copy, revision = _sync_working_copy(runner, repository_url, relative_path, username, password)
            working_copies[source_ref] = working_copy
            revisions[source_ref] = revision
        index_path = settings.knowledge_root / "published" / "svn-index.sqlite3"
        previous = _load_manifest(index_path)
        manifest, changes = _build_manifest(repository_urls, revisions, working_copies, previous)
        dimensions = _publish_vector_index(
            manifest,
            previous,
            working_copies,
            EmbeddingClient(embedding_base_url, embedding_model, embedding_api_key),
        )
    except Exception as exc:
        db = SessionLocal()
        try:
            source = db.get(SvnKnowledgeSource, source_id)
            if source:
                source.sync_status = "failed"
                source.last_error = str(redact(str(exc)))[:1000]
                db.commit()
        finally:
            db.close()
        raise

    db = SessionLocal()
    try:
        source = db.get(SvnKnowledgeSource, source_id)
        if source:
            source.sync_status = "succeeded"
            source.last_success_at = beijing_now()
            source.last_revisions = revisions
            source.failed_file_count = len(manifest.get("failed_files", {}))
            source.file_count = len(manifest["files"]) - source.failed_file_count
            source.last_changes = changes
            source.embedding_dimensions = dimensions
            source.last_error = None
            db.commit()
    finally:
        db.close()
