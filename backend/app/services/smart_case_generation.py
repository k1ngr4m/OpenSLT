from __future__ import annotations

import hashlib
import os
import typing
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.core.database import SessionLocal
from app.core.logging import redact
from app.core.security import decrypt_secret
from app.core.time import beijing_now
from app.models import SmartCaseGeneration, SvnKnowledgeSource
from app.services.embedding import EmbeddingClient
from app.services.llm import LlmClient, generate_cases
from app.services.svn_knowledge import get_indexed_document, published_index_matches, search_vector_index


def _safe_excel(value: typing.Any) -> str:
    text = str(value or "")
    return "'" + text if text.startswith(("=", "+", "-", "@")) else text


def build_workbook(path: Path, generation: SmartCaseGeneration, cases: typing.Sequence[typing.Mapping[str, typing.Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "测试用例"
    headers = ["用例编号", "需求编号", "需求名称", "用例名称", "前置条件", "测试步骤", "预期结果", "用例类型", "优先级", "状态", "来源", "备注"]
    sheet.append(headers)
    for index, case in enumerate(cases, 1):
        sheet.append([
            "TC-%04d" % index,
            generation.requirement_no or "",
            generation.requirement_name,
            case["title"],
            "\n".join(case["preconditions"]),
            "\n".join("%d. %s" % (i, value) for i, value in enumerate(case["steps"], 1)),
            "\n".join("%d. %s" % (i, value) for i, value in enumerate(case["expected_results"], 1)),
            case["case_type"], case["priority"], "草稿待复核", generation.requirement_path, "",
        ])
    fill = PatternFill("solid", fgColor="14545A")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.value = _safe_excel(cell.value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [14, 18, 24, 34, 28, 48, 48, 12, 10, 14, 42, 24]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    notes = workbook.create_sheet("生成说明")
    notes.append(["项目", "内容"])
    notes.append(["生成模型", generation.llm_model])
    notes.append(["生成时间", beijing_now().isoformat()])
    notes.append(["选中需求", generation.requirement_path])
    notes.append(["需求 revision", generation.requirement_revision])
    notes.append(["复核要求", "本文件为 AI 生成草稿，执行前必须由测试人员复核。"])
    for item in generation.referenced_sources:
        notes.append(["参考来源", "%s（r%s）" % (item["source_path"], item["revision"])])
    for cell in notes[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
    notes.column_dimensions["A"].width = 18
    notes.column_dimensions["B"].width = 90
    notes.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(path.name + "." + uuid4().hex + ".tmp")
    try:
        workbook.save(temporary)
        os.replace(str(temporary), str(path))
        path.chmod(0o600)
    finally:
        workbook.close()
        temporary.unlink(missing_ok=True)


def execute_smart_case_generation(generation_id: int) -> None:
    db = SessionLocal()
    try:
        generation = db.get(SmartCaseGeneration, generation_id)
        source = db.query(SvnKnowledgeSource).order_by(SvnKnowledgeSource.id).first()
        if generation is None or source is None:
            raise RuntimeError("智能用例生成任务或知识源不存在")
        if generation.status == "succeeded":
            return
        if not published_index_matches(source):
            raise RuntimeError("知识索引已变化，请重新选择需求")
        generation.status = "running"
        generation.error = None
        db.commit()
        selected = get_indexed_document(generation.requirement_path)
        query = "%s %s %s" % (generation.requirement_no or "", generation.requirement_name, selected["content"][:1500])
        embedding = EmbeddingClient(source.embedding_base_url, source.embedding_model, decrypt_secret(source.encrypted_embedding_api_key))
        vector = embedding.embed([query])[0]
        hits = search_vector_index(query, vector, 8)
        references = [{"source_path": selected["source_path"], "revision": selected["revision"], "content": selected["content"][:12000]}]
        seen = {selected["source_path"]}
        for hit in hits:
            if hit["source_path"] in seen:
                continue
            seen.add(hit["source_path"])
            references.append({"source_path": hit["source_path"], "revision": hit["revision"], "content": hit["snippet"][:1500]})
        generation.referenced_sources = [{"source_path": item["source_path"], "revision": item["revision"]} for item in references]
        cases = generate_cases(
            LlmClient(source.llm_base_url, source.llm_model, decrypt_secret(source.encrypted_llm_api_key)),
            {"requirement_no": generation.requirement_no or "", "requirement_name": generation.requirement_name, "source_path": generation.requirement_path, "revision": generation.requirement_revision},
            references,
        )
        path = settings.artifact_root / "smart-cases" / ("generation-%s.xlsx" % generation.id)
        build_workbook(path, generation, cases)
        digest = hashlib.sha256(path.read_bytes()).hexdigest()
        generation.result_cases = list(cases)
        generation.case_count = len(cases)
        generation.artifact_path = str(path)
        generation.artifact_size = path.stat().st_size
        generation.artifact_checksum = digest
        generation.status = "succeeded"
        db.commit()
    except Exception as exc:
        db.rollback()
        generation = db.get(SmartCaseGeneration, generation_id)
        if generation:
            generation.status = "failed"
            generation.error = str(redact(str(exc)))[:1000]
            db.commit()
        raise
    finally:
        db.close()
