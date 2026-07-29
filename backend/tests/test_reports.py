from __future__ import annotations

import hashlib
from pathlib import Path

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

from app.core.config import settings
from app.core.database import SessionLocal
from app.core.time import beijing_now
from app.models import (
    Artifact,
    ConfigurationCaptureItem,
    ConfigurationCaptureSnapshot,
    RunStep,
    ScenarioWorkflowNode,
    ScenarioWorkflowVersion,
    TestRun as RunModel,
    TestScenario as ScenarioModel,
)
from app.services.reports import build_report_document, generate_reports
from app.services.workflow_contracts import _archive_order_config
from app.services.workflow_core import WorkflowError
from conftest import create_plan_scenario, create_resource


RAW_XML = '''<?xml version="1.0" encoding="utf-8"?>
<tcp>
  <account id="main" disp="账户"><password disp="密码" value="p&amp;&lt;secret&gt;" /></account>
  <note disp="超长配置" value="''' + ("中文配置" * 80) + '''" />
</tcp>'''


def _statistics_result(source: str, offset: int) -> dict:
    return {
        "source_path": source,
        "source_file": Path(source).name,
        "unit": "ns",
        "sample_count": 100 + offset,
        "excluded_counts": {"above_limit": offset, "negative": 1, "invalid": 2},
        "metrics": [
            {"key": "p99", "label": "P99", "value": 901.23456 + offset},
            {"key": "average", "label": "平均值", "value": 123.45678 + offset},
            {"key": "maximum", "label": "最大值", "value": 999 + offset},
            {"key": "minimum", "label": "最小值", "value": 12 + offset},
            {"key": "median", "label": "中位数", "value": 100 + offset},
            {"key": "stddev", "label": "标准差", "value": 8.7654 + offset},
            {"key": "custom_jitter", "label": "抖动", "value": 3.14159 + offset},
        ],
    }


def _create_report_run(client, admin_headers, tmp_path: Path):
    resource = create_resource(client, admin_headers, "REM-报告")
    plan, scenario_data = create_plan_scenario(
        client, admin_headers, resource_ids=[resource["id"]]
    )
    with SessionLocal() as db:
        scenario = db.get(ScenarioModel, scenario_data["id"])
        version = db.get(ScenarioWorkflowVersion, scenario.draft_workflow_version_id)
        node_types = [
            "server_config", "database_config", "order_preparation",
            "data_statistics", "data_statistics", "report_generation",
        ]
        nodes = []
        for position, node_type in enumerate(node_types, 1):
            node = ScenarioWorkflowNode(
                workflow_version=version,
                node_key=f"node-{position}",
                position=position,
                node_type=node_type,
                name=f"{node_type}-{position}",
                config={},
            )
            db.add(node)
            nodes.append(node)
        db.flush()
        run = RunModel(
            run_number="R-REPORT-001",
            plan_id=plan["id"],
            scenario_id=scenario.id,
            workflow_version_id=version.id,
            business_code="fut_mm",
            status="completed",
            progress=100,
            resource_ids=[resource["id"]],
            config_snapshot={
                "plan": {"name": "期货测速方案"},
                "scenario": {"name": "发单延迟"},
                "workflow": {"version_id": version.id},
                "resources": [{
                    "id": resource["id"], "name": "REM-报告",
                    "type": "rem", "host": "10.10.0.8",
                }],
            },
            trace_id="report-test",
            created_by=1,
            started_at=beijing_now(),
            finished_at=beijing_now(),
        )
        steps = []
        for position, node in enumerate(nodes, 1):
            summary = {}
            if node.node_type == "data_statistics":
                summary = {
                    "statistics_script": {"filename": f"statistics-{position}.py"},
                    "statistics_results": [
                        _statistics_result(f"结果目录/来源-{index}.csv", position + index)
                        for index in range(1, 9)
                    ],
                }
            step = RunStep(
                workflow_node_id=node.id,
                code=node.node_key,
                name=node.name,
                node_type=node.node_type,
                config_snapshot={},
                result_summary=summary,
                position=position,
                status="succeeded",
                progress=100,
                started_at=beijing_now(),
                finished_at=beijing_now(),
                duration_ms=position * 10,
            )
            steps.append(step)
        run.steps = steps
        db.add(run)
        db.flush()

        for step, source_type, database_name, label, value in (
            (steps[0], "server", None, "CPU 型号", "兆芯 KX-7000"),
            (steps[1], "database", "alpha_config", "trade.timeout_ns", "200000"),
        ):
            snapshot = ConfigurationCaptureSnapshot(
                scenario_id=scenario.id,
                workflow_version_id=version.id,
                workflow_node_id=step.workflow_node_id,
                run_id=run.id,
                run_step_id=step.id,
                scope="run",
                source_type=source_type,
                resource_id=resource["id"],
                database_name=database_name,
                status="succeeded",
                created_by=1,
                finished_at=beijing_now(),
            )
            snapshot.items = [ConfigurationCaptureItem(
                item_key=label,
                item_label=label,
                value_text=value,
                source_reference="runtime",
                raw_output=value,
                exit_code=0,
                status="succeeded",
            )]
            db.add(snapshot)

        xml_path = tmp_path / "validated-order.xml"
        xml_path.write_text(RAW_XML, encoding="utf-8")
        xml_data = xml_path.read_bytes()
        db.add(Artifact(
            run_id=run.id,
            step_id=steps[2].id,
            artifact_type="order_config_xml",
            name="validated-order.xml",
            path=str(xml_path),
            content_type="application/xml; charset=utf-8",
            size=len(xml_data),
            checksum=hashlib.sha256(xml_data).hexdigest(),
            is_immutable=True,
        ))
        ordinary_path = settings.artifact_root / "test-inputs" / "ordinary.txt"
        ordinary_path.parent.mkdir(parents=True, exist_ok=True)
        ordinary_path.write_text("ordinary", encoding="utf-8")
        db.add(Artifact(
            run_id=run.id,
            step_id=steps[3].id,
            artifact_type="statistics_result_json",
            name=ordinary_path.name,
            path=str(ordinary_path),
            content_type="text/plain",
            size=ordinary_path.stat().st_size,
            checksum=hashlib.sha256(ordinary_path.read_bytes()).hexdigest(),
            is_immutable=True,
        ))
        db.commit()
        return run.id, steps[-1].id


def test_report_document_metric_matrix_and_missing_sections(client, admin_headers, tmp_path):
    run_id, report_step_id = _create_report_run(client, admin_headers, tmp_path)
    with SessionLocal() as db:
        run = db.get(RunModel, run_id)
        report_step = db.get(RunStep, report_step_id)
        document = build_report_document(db, run, 1, beijing_now(), report_step)
        assert document["missing_sections"] == []
        assert document["servers"][0]["host"] == "10.10.0.8"
        assert document["databases"][0]["database_name"] == "alpha_config"
        assert document["orders"][0]["raw_xml"] == RAW_XML
        assert len(document["statistics"]) == 2
        assert document["statistics"][0]["columns"][0] == "结果目录/来源-1.csv"
        labels = [row["label"] for row in document["statistics"][0]["rows"]]
        assert labels[:6] == ["Avg", "Max", "Min", "Md", "Std", "Cnt"]
        assert labels[-1] == "抖动"
        avg = document["statistics"][0]["rows"][0]["values"][0]
        assert avg == "128.457"

        run.steps[0].result_summary = {}
        empty = build_report_document(db, run, 1, beijing_now(), run.steps[0])
        assert empty["missing_sections"] == ["servers", "databases", "orders", "statistics"]


def test_report_versions_are_immutable_and_render_all_formats(
    client, admin_headers, tmp_path, monkeypatch
):
    monkeypatch.setattr(settings, "artifact_root", tmp_path / "artifacts")
    run_id, report_step_id = _create_report_run(client, admin_headers, tmp_path)
    with SessionLocal() as db:
        run = db.get(RunModel, run_id)
        report_step = db.get(RunStep, report_step_id)
        first = generate_reports(db, run, step=report_step, reason="workflow_node")
        db.commit()
        first_artifacts = [db.get(Artifact, artifact_id) for artifact_id in first["artifact_ids"]]
        first_checksums = {item.id: item.checksum for item in first_artifacts}
        first_paths = {item.artifact_type: Path(item.path) for item in first_artifacts}

        second = generate_reports(db, run, step=report_step, reason="manual")
        db.commit()
        assert first["report_version"] == 1
        assert second["report_version"] == 2
        assert report_step.result_summary["report_version"] == 2
        assert db.query(Artifact).filter(
            Artifact.run_id == run.id,
            Artifact.artifact_type.in_(("web_report", "excel_report", "pdf_report")),
        ).count() == 6
        for artifact_id, checksum in first_checksums.items():
            artifact = db.get(Artifact, artifact_id)
            assert artifact.checksum == checksum
            assert artifact.is_immutable is True
            assert Path(artifact.path).is_file()

        html_text = first_paths["web_report"].read_text(encoding="utf-8")
        assert "&lt;password" in html_text
        assert "p&amp;amp;&amp;lt;secret&amp;gt;" in html_text
        workbook = load_workbook(first_paths["excel_report"], read_only=True)
        assert {"运行摘要", "服务器配置", "数据库配置", "发单配置", "步骤时间线", "XML原文-1"}.issubset(workbook.sheetnames)
        assert len([name for name in workbook.sheetnames if name.startswith("指标-")]) == 2
        assert "<password" in "\n".join(
            str(row[0]) for row in workbook["XML原文-1"].iter_rows(values_only=True) if row[0]
        )
        reader = PdfReader(str(first_paths["pdf_report"]))
        assert len(reader.pages) >= 4
        sizes = {(round(float(page.mediabox.width)), round(float(page.mediabox.height))) for page in reader.pages}
        assert any(width < height for width, height in sizes)
        assert any(width > height for width, height in sizes)
        extracted = "\n".join(page.extract_text() or "" for page in reader.pages)
        assert "OpenSLT" in extracted
        assert "兆芯" in extracted
        assert extracted.count("指标") >= 4

    visitor = client.post("/api/v1/users", headers=admin_headers, json={
        "username": "report-viewer", "display_name": "报告访客",
        "password": "viewer-password", "role": "visitor",
    })
    assert visitor.status_code == 201, visitor.text
    token = client.post("/api/v1/auth/login", json={
        "username": "report-viewer", "password": "viewer-password",
    }).json()["access_token"]
    visitor_headers = {"Authorization": f"Bearer {token}"}
    assert client.get(
        f"/api/v1/artifacts/{first['artifact_ids'][0]}/download", headers=visitor_headers
    ).status_code == 403
    with SessionLocal() as db:
        xml_id = db.query(Artifact.id).filter(
            Artifact.run_id == run_id, Artifact.artifact_type == "order_config_xml"
        ).scalar()
        ordinary_id = db.query(Artifact.id).filter(
            Artifact.run_id == run_id, Artifact.artifact_type == "statistics_result_json"
        ).scalar()
    assert client.get(f"/api/v1/artifacts/{xml_id}/download", headers=visitor_headers).status_code == 403
    assert client.get(f"/api/v1/artifacts/{ordinary_id}/download", headers=visitor_headers).status_code == 200
    assert client.get(
        f"/api/v1/artifacts/{first['artifact_ids'][0]}/download", headers=admin_headers
    ).status_code == 200


def test_order_xml_archive_is_reused_and_detects_changes(
    client, admin_headers, tmp_path, monkeypatch
):
    monkeypatch.setattr(settings, "artifact_root", tmp_path / "artifacts")
    run_id, _ = _create_report_run(client, admin_headers, tmp_path)
    checksum = hashlib.sha256(RAW_XML.encode("utf-8")).hexdigest()
    with SessionLocal() as db:
        run = db.get(RunModel, run_id)
        step = next(item for item in run.steps if item.node_type == "order_preparation")
        db.query(Artifact).filter(
            Artifact.run_id == run.id, Artifact.artifact_type == "order_config_xml"
        ).delete(synchronize_session=False)
        db.flush()
        first = _archive_order_config(db, run, step, "remote.xml", RAW_XML, checksum)
        second = _archive_order_config(db, run, step, "remote.xml", RAW_XML, checksum)
        assert first.id == second.id
        assert Path(first.path).read_text(encoding="utf-8") == RAW_XML
        Path(first.path).write_text("changed", encoding="utf-8")
        with pytest.raises(WorkflowError) as exc:
            _archive_order_config(db, run, step, "remote.xml", RAW_XML, checksum)
        assert exc.value.code == "ORDER_CONFIG_ARCHIVE_CHANGED"


def test_workflow_verdict_updates_and_manual_regeneration_create_versions(
    client, admin_headers, tmp_path, monkeypatch
):
    monkeypatch.setattr(settings, "artifact_root", tmp_path / "artifacts")
    run_id, report_step_id = _create_report_run(client, admin_headers, tmp_path)
    endpoint = f"/api/v1/runs/{run_id}/verdict"
    first = client.post(endpoint, headers=admin_headers, json={
        "final_result": "passed", "issue_description": "", "notes": "首次复核",
    })
    assert first.status_code == 200, first.text
    second = client.post(endpoint, headers=admin_headers, json={
        "final_result": "conditional", "issue_description": "尾延迟偏高", "notes": "复测",
    })
    assert second.status_code == 200, second.text
    regenerated = client.post(f"/api/v1/runs/{run_id}/reports", headers=admin_headers)
    assert regenerated.status_code == 200, regenerated.text
    assert {item["name"] for item in regenerated.json()} == {
        "report-v003.html", "report-v003.xlsx", "report-v003.pdf",
    }
    with SessionLocal() as db:
        report_step = db.get(RunStep, report_step_id)
        assert report_step.result_summary["report_version"] == 3
        report_artifacts = db.query(Artifact).filter(
            Artifact.run_id == run_id,
            Artifact.artifact_type.in_(("web_report", "excel_report", "pdf_report")),
        ).all()
        assert len(report_artifacts) == 9
        v1_html = next(item for item in report_artifacts if item.name == "report-v001.html")
        v2_html = next(item for item in report_artifacts if item.name == "report-v002.html")
        assert "首次复核" in Path(v1_html.path).read_text(encoding="utf-8")
        assert "尾延迟偏高" in Path(v2_html.path).read_text(encoding="utf-8")


def test_legacy_verdict_still_completes_fixed_step_run(
    client, admin_headers, tmp_path, monkeypatch
):
    monkeypatch.setattr(settings, "artifact_root", tmp_path / "artifacts")
    plan, scenario = create_plan_scenario(client, admin_headers)
    with SessionLocal() as db:
        run = RunModel(
            run_number="R-LEGACY-REPORT",
            plan_id=plan["id"],
            scenario_id=scenario["id"],
            workflow_version_id=None,
            business_code="fut_mm",
            status="awaiting_review",
            progress=90,
            resource_ids=[],
            config_snapshot={"plan": {"name": "旧方案"}, "scenario": {"name": "旧场景"}},
            trace_id="legacy-report-test",
            created_by=1,
        )
        run.steps = [
            RunStep(
                code="manual_review", name="人工复核", node_type="legacy", position=1,
                status="waiting", progress=100, config_snapshot={}, result_summary={},
            ),
            RunStep(
                code="reporting", name="生成报告", node_type="legacy", position=2,
                status="pending", progress=0, config_snapshot={}, result_summary={},
            ),
        ]
        db.add(run)
        db.commit()
        run_id = run.id

    response = client.post(f"/api/v1/runs/{run_id}/verdict", headers=admin_headers, json={
        "final_result": "passed", "issue_description": "", "notes": "旧版兼容",
    })
    assert response.status_code == 200, response.text
    detail = client.get(f"/api/v1/runs/{run_id}", headers=admin_headers).json()
    assert detail["status"] == "completed"
    assert [step["status"] for step in detail["steps"]] == ["succeeded", "succeeded"]
    assert {item["name"] for item in detail["artifacts"]} == {
        "report-v001.html", "report-v001.xlsx", "report-v001.pdf",
    }
